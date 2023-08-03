import configparser
import os
import sys
import time
import pymysql
import requests
import json
import devConfig
import prdConfig
from openpyxl import Workbook
import datetime

conn = ''
try:
    # 校验日期格式是否为YYYY-MM-dd
    def check_date(d):

        try:
            time.strptime(d, "%Y-%m-%d")
            return True
        except Exception:
            return False


    # 读取配置文件
    def read_config(fileName):
        # path = resource_path(fileName)
        # print(path)
        config = configparser.RawConfigParser()
        config.read(fileName)
        return config


    def resource_path(relative_path):
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)


    # 登陆接口
    def login(userName, password):
        loginUrl = '/gate/login/mpLogin?username={userName}&password={password}'.format(userName=userName,
                                                                                        password=password)
        # 请求接口
        res = requests.post(apiHost + loginUrl)
        # 对返回的内容进行编码
        content = res.content.decode('utf-8')
        # 将json字符串反序列化
        tokenJson = json.loads(content)
        return tokenJson


    # 获取投料卡
    def dosageAllMatList(token, retailCodes):
        dosageUrl = '/order-service/dosageAllMatList?addFlag=Y'
        x_header = {
            'Content-Type': 'application/json; charset=utf-8',
            'X-Auth-Token': token

        }
        body = json.dumps(retailCodes)
        res = requests.post(apiHost + dosageUrl, data=body, headers=x_header, timeout=180)
        # 对返回的内容进行编码
        content = res.content.decode('utf-8')
        # 将json字符串反序列化
        listJson = json.loads(content)
        return listJson['success']


    iniName = input("请输入调用环境,默认为生产环境，如需调用开发环境请输入dev，反之回车即可>>>>>:") or 'prd'
    print('当前调用环境为：' + iniName)
    if iniName == 'prd':
        configInfo = prdConfig
    else:
        configInfo = devConfig
    # config = read_config(iniName)
    # 请求域名
    apiHost = configInfo.config.apihost
    userName = input("请输入门店进销存登陆账号：") or configInfo.config.default_user_name if iniName == 'zjb' else input(
        "请输入门店进销存登陆账号：")
    password = input(
        "请输入门店进销存登陆密码：") or configInfo.config.default_user_password if iniName == 'zjb' else input(
        "请输入门店进销存登陆密码：")

    loginJson = login(userName, password)
    if loginJson['success']:
        userInfo = loginJson['obj']
        token = userInfo['token']
    else:
        print('登陆失败，失败原因：' + loginJson['msg'])
        exit(0)
    print('登陆成功')
    retailCode = input("请输入要查询的门店编码，多个门店编码之间逗号分割,如20133,10001>>>>>:").replace(" ", "").replace(
        "\n",
        "").strip(
        ',').split(',')
    # print(retailCode)
    dosageFlag = input("是否重新拉取MDC投料卡，默认不拉取，如需拉取请输入是，反之回车即可>>>>>:") or '否'
    if dosageFlag == '是':
        print('开始拉取MDC投料卡信息')
        if dosageAllMatList(token, retailCode):
            print('获取投料卡成功,开始查询报货单')
        else:
            print('投料卡获取失败，请联系赵俊彪')
            exit(0)
    now = datetime.datetime.now().strftime('%Y-%m-%d')
    while True:
        startDate = input("请输入要查询的报货单开始日期，如:2023-04-12>>>>>:").rstrip()
        if check_date(startDate):
            break
        else:
            print('开始日期格式不对，请重新输入')
            continue
    while True:

        endDate = input(
            "请输入要查询的报货单结束日期，如:" + now + ">>>>>:").rstrip() or now if iniName == 'zjb' else input("请输入要查询的报货单结束日期，如:" + now + ">>>>>:").rstrip()
        if check_date(endDate):
            break
        else:
            print('结束日期格式不对，请重新输入')
            continue

    org_code_list = ','.join('%s' % id for id in retailCode)

    # 一、获取数据库连接
    conn = pymysql.connect(
        host=configInfo.config.host,
        user=configInfo.config.user,
        password=configInfo.config.password,
        db=configInfo.config.db_order,
        charset='utf8',
    )
    curs = conn.cursor()
    print("开始连接数据库：成功!")

    sqls = [
        'SELECT retail_code,retail_name,require_reason_code,product_code,product_name,CONCAT(retail_code,product_code) \'门店+物资\' FROM ord_mdc_require_product WHERE RETAIL_CODE IN ({retailCode})',
        'SELECT re.retail_code , red.product_code ,red.product_name ,red.product_spec ,red.unit_name ,SUM(red.require_qty) ,red.unit_price ,SUM(red.require_amount),re.require_status,CONCAT(red.retail_code,red.product_code) \'门店+物资\' FROM ord_retail_dept_require re LEFT JOIN ord_retail_dept_require_detail red ON re.id = red.dept_require_id WHERE DATE_FORMAT( re.require_date, \'%Y-%m-%d\' ) BETWEEN \'{startDate}\' AND \'{endDate}\' AND re.require_status != \"8BE_CANCLE\" AND  red.category_code IN (\'1\',\'2\',\'3\') AND re.retail_code IN ({retailCode})  AND re.code NOT LIKE \'PT%\' GROUP BY re.retail_code,red.product_code',
        'SELECT product_code,product_name, tag_code,struct_attribute, is_enabled FROM com_data_product WHERE struct_attribute = \'0\' OR (tag_code <> - 1 AND tag_code IS NOT NULL ) AND is_enabled = \'Y\'',
        'SELECT product_code,product_name,provider_code,provider_name,org_code,org_name,CONCAT(org_code,product_code) \'门店+物资\' FROM com_data_product_org WHERE org_code IN ({retailCode}) AND relation_status = \'Y\'']

    # 创建excel
    w = Workbook()
    initSheet = w.active
    w.remove(initSheet)
    for sqlI in range(len(sqls)):
        sheetName = 'sheet'
        title = 'a,b,c,d,e,f'
        if sqlI == 0:
            sql = sqls[sqlI].format(retailCode=org_code_list)
            title = "retail_code,retail_name,require_reason_code,product_code,product_name,门店+物资"
            sheetName = '投料卡'
        if sqlI == 1:
            sql = sqls[sqlI].format(retailCode=org_code_list, startDate=startDate, endDate=endDate)
            title = "门店编码,物资编码,物资名称,规格,单位,报货数量,单价,金额,状态,门店+物资,标签+母件核对,投料卡核对,组织物资供应商关系核对"
            sheetName = '报货单'

        if sqlI == 2 or sqlI == 3:
            conn.close()
            conn = pymysql.connect(
                host=configInfo.config.host,
                user=configInfo.config.user,
                password=configInfo.config.password,
                db=configInfo.config.db_data,
                charset='utf8',
            )
            curs = conn.cursor()
            if sqlI == 2:
                sql = sqls[sqlI]
                title = "product_code,product_name,tag_code,struct_attribute,is_enabled"
                sheetName = '标签+母件'
            if sqlI == 3:
                sql = sqls[sqlI].format(retailCode=org_code_list)
                title = "product_code,product_name,provider_code,provider_name,org_code,org_name,门店+物资"
                sheetName = '组织物资供应商关系'
        # print(sql)

        curs.execute(sql)
        if sqlI == 0:
            print('查询投料卡完成!')
        if sqlI == 1:
            print('查询报货单完成!')
        if sqlI == 2:
            print('查询标签+母件完成!')
        if sqlI == 3:
            print('查询组织物资供应商关系完成!')
        rows = curs.fetchall()  # 获取所有数据
        # print(rows)

        ws = w.create_sheet(title=sheetName)
        # 三、把数据写入excel

        columnName = []  # 定义一个excel列名列表
        for c in title.split(","):  # 循环遍历数据库中返回的列名元组信息
            columnName.append(c)  # 获取列名的字符串，并添加到列名列表
        for i in range(1, len(columnName) + 1):  # 循环遍历列名列表，并将列名写入sheet对象的首行
            ws.cell(row=1, column=i, value=columnName[i - 1])  # 将列名写入表格
        for results in rows:  # 循环遍历接收到的数据列表
            ws.append(results)  # 将对应的行列坐标的数据写入excel

    conn.close()
    print('关闭数据库连接：成功!')
    print('开始生成excel文档!')
    # 四、数据保存到本地
    filName = '投料卡报货{}.xlsx'.format(time.strftime('%Y%m%d%H%M%S', time.localtime()))
    w.save(filName)
    print('excel文档生成完毕，开始下载!')
    print("下载成功!")
except Exception as msg:
    print(msg.__traceback__.tb_lineno)
    print("导出错误，异常信息：", msg)
finally:
    if conn != '' and conn.open == 1:
        conn.close()
